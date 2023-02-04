import sys



sys.dont_write_bytecode = True
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
import django
django.setup()
from extract.act import get_act_details
from utils import to_excel
from extract.page_extractor import get_page
from save import base_dir, skipped_pages_list_dir, files_list_dir
from extract.act_list import get_act_list_single_page
from openpyxl import load_workbook

from save.act import already_added

from db.models import Law




# Django specific settings



# Import your models for use in your script





def crawl():
    wb = load_workbook(skipped_pages_list_dir)
    ws = wb.worksheets[0]

    wb2 = load_workbook(files_list_dir)
    ws2 = wb2.worksheets[0]
    page_file_dir = base_dir + f"/fetched_pages.txt"
    if not os.path.isfile(page_file_dir):
        f = open(page_file_dir, "w")
        f.write(str(1) + "\n")
        f.close()
    file = open(page_file_dir, "r")
    page = file.readline()
    file.close()
    page = int(page)
    count = 0
    stored_exception = None

    while True:
        acts = []
        try:
            print(f'fetching page {page} ...')
            page_content = get_page(page)
            acts = get_act_list_single_page(page_content)
        except KeyboardInterrupt or SystemExit:
            stored_exception = sys.exc_info()
        except:
            ws.append([page])
            wb.save(skipped_pages_list_dir)
            print(f'error fetching page {page}.')
            page += 1
            continue
        for index_, act in enumerate(acts):
            try:
                record = get_act_details(act)
                if already_added(str(record['id'])):
                    print(f'act {record} already added')
                    continue
                    # print(record['id'])
                    # print("inside")
                Law.objects.create(name=record['name'],pid=record['id'],data=record['date'],approve=record['approval'])
                # ws2.append(
                #     [record['id'], record['name'], record['date'], record['approval']])
                # wb2.save(files_list_dir)
                count += 1
            except KeyboardInterrupt or SystemExit:
                stored_exception = sys.exc_info()
        if stored_exception or len(acts) == 0:
            break
        page += 1
        try:
            f = open(page_file_dir, "w")
            f.write(str(page) + "\n")
            f.close()
        except KeyboardInterrupt or SystemExit:
            stored_exception = sys.exc_info()
        print(f'total processed acts :{count}')
        # if page % 10 == 0:

        #     break


    if stored_exception:
        print("Either max act count limit reached or user stopped the process!")
    else:
        print(f"done")



# to_excel()
crawl()




